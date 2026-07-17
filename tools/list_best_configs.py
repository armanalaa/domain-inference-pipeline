"""
list_best_configs.py
====================
Reads the first row (best Q) from each dataset's tune_params_results.xlsx,
reads all Q > 0.3 configurations, and writes a single summary Excel file
to the results folder.

USAGE
-----
  python tools/list_best_configs.py
  python tools/list_best_configs.py --datasets Sakila Northwind Chinook
  python tools/list_best_configs.py --output results/best_configs.xlsx
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# =============================================================================
# Configuration
# =============================================================================

PROJECT_ROOT = Path(__file__).resolve().parents[1]
PIPELINE_DIR = PROJECT_ROOT / "pipeline"
if str(PIPELINE_DIR) not in sys.path:
    sys.path.insert(0, str(PIPELINE_DIR))

from path_utils import DATALAKES_ROOT

RUN_PATTERN = re.compile(r"tA([\d.]+)_tT([\d.]+)_r([\d.]+)")


def discover_datasets(root: Path) -> list[str]:
    """Find every subfolder of root that contains a tune_params_results.xlsx.

    This replaces the old hardcoded KNOWN_DATASETS list so newly added
    datasets (e.g. Mondial, Lahman, IMDB JOB) are picked up automatically
    without editing this script.
    """
    found = []
    for child in sorted(root.iterdir()):
        if not child.is_dir():
            continue
        xlsx_path = child / "ccm_output" / "tune_params_results.xlsx"
        if xlsx_path.exists():
            found.append(child.name)
    return found

# =============================================================================
# Read dataset stats from schema.json
# =============================================================================

def read_dataset_stats(dataset_dir: Path) -> dict:
    """Read total tables and columns from schema.json."""
    schema_path = dataset_dir / "schema.json"
    total_tables  = None
    total_columns = 0

    # Tables and columns from schema.json
    if schema_path.exists():
        try:
            data   = json.loads(schema_path.read_text(encoding="utf-8", errors="ignore"))
            tables = data.get("tables", {})
            if isinstance(tables, list):
                tables = {t.get("name", str(i)): t for i, t in enumerate(tables)}
            total_tables = len(tables)
            for tdata in tables.values():
                if isinstance(tdata, dict):
                    total_columns += len(tdata.get("columns", []))
        except Exception as e:
            print(f"  [WARN] schema.json unreadable for {dataset_dir.name}: {e}")

    return {
        "total_tables":  total_tables,
        "total_columns": total_columns if total_columns > 0 else None,
        "csv_tables": count_csv_tables(dataset_dir),
    }

# =============================================================================

def count_csv_tables(dataset_dir: Path) -> int | None:
    """Count CSV files, treating each CSV as one source table."""
    csv_dir = dataset_dir / "csv"
    if not csv_dir.exists():
        return None
    csv_files = [
        p for p in csv_dir.iterdir()
        if p.is_file() and p.suffix.lower() == ".csv"
    ]
    return len(csv_files)


def read_domains_json(dataset_dir: Path, run_tag: str) -> list[dict] | None:
    """Read step5_domains.json for a run."""
    domains_path = dataset_dir / "ccm_output" / run_tag / "step5_domains.json"
    if not domains_path.exists():
        return None
    try:
        data = json.loads(domains_path.read_text(encoding="utf-8", errors="ignore"))
        return data if isinstance(data, list) else None
    except Exception:
        return None


def sum_domain_tables(dataset_dir: Path, run_tag: str) -> int | None:
    """Sum the number of tables assigned across all domains in a run."""
    data = read_domains_json(dataset_dir, run_tag)
    if data is None:
        return None

    total = 0
    for domain in data:
        if not isinstance(domain, dict):
            continue
        tables = domain.get("tables")
        if isinstance(tables, list):
            total += len(tables)
    return total


def table_count_check(domain_tables: int | None, schema_tables: int | None, csv_tables: int | None) -> str:
    """Compare domain table sum against schema and CSV table counts."""
    expected = [v for v in (schema_tables, csv_tables) if v is not None]
    if domain_tables is None or not expected:
        return "unknown"
    return "OK" if all(domain_tables == v for v in expected) else "MISMATCH"

def read_best_from_xlsx(dataset_dir: Path) -> dict | None:
    """Read the first data row (highest Q) from tune_params_results.xlsx."""
    xlsx_path = dataset_dir / "ccm_output" / "tune_params_results.xlsx"
    if not xlsx_path.exists():
        return None
    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        ws = wb.active
        # Row 1 = title, Row 2 = headers, Row 3 = best run (sorted by Q desc)
        row = list(ws.iter_rows(min_row=3, max_row=3, values_only=True))[0]
        if row[0] is None:
            return None
        table_domain_ratio = (
            row[4] / row[6]
            if row[4] is not None and row[6] not in (None, 0)
            else None
        )
        return {
            "theta_a":    row[0],
            "theta_t":    row[1],
            "resolution": row[2],
            "run_tag":    row[3],
            "n_tables":   row[4],
            "n_edges":    row[5],
            "n_domains":  row[6],
            "Q":          row[7],
            "status":     row[8],
            "elapsed_s":  row[9],
            "table_domain_ratio": table_domain_ratio,
            "single_table_domains": count_single_table_domains(dataset_dir, str(row[3])),
            "domain_table_sum": sum_domain_tables(dataset_dir, str(row[3])),
        }
    except Exception as e:
        print(f"  [WARN] Could not read {xlsx_path}: {e}")
        return None


def read_valid_configs_from_xlsx(dataset_dir: Path, min_q: float = 0.3) -> list[dict]:
    """Read every configuration whose Q is above min_q from tune_params_results.xlsx."""
    xlsx_path = dataset_dir / "ccm_output" / "tune_params_results.xlsx"
    if not xlsx_path.exists():
        return []
    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        ws = wb.active
        configs = []
        for row in ws.iter_rows(min_row=3, values_only=True):
            if not row or row[0] is None:
                continue

            q = row[7]
            if q is None or q <= min_q:
                continue

            n_tables = row[4]
            n_domains = row[6]
            table_domain_ratio = (
                n_tables / n_domains
                if n_tables is not None and n_domains not in (None, 0)
                else None
            )
            configs.append({
                "theta_a": row[0],
                "theta_t": row[1],
                "resolution": row[2],
                "run_tag": row[3],
                "n_tables": n_tables,
                "n_edges": row[5],
                "n_domains": n_domains,
                "Q": q,
                "status": row[8],
                "elapsed_s": row[9],
                "table_domain_ratio": table_domain_ratio,
                "single_table_domains": count_single_table_domains(dataset_dir, str(row[3])),
                "domain_table_sum": sum_domain_tables(dataset_dir, str(row[3])),
            })

        return sorted(
            configs,
            key=lambda c: (
                c["table_domain_ratio"] is not None,
                c["table_domain_ratio"] or 0,
                c["Q"] or 0,
            ),
            reverse=True,
        )
    except Exception as e:
        print(f"  [WARN] Could not read all configs from {xlsx_path}: {e}")
        return []

# =============================================================================
# Read domain names from step5_domains.json of the best run folder
# =============================================================================

def read_domain_names(dataset_dir: Path, run_tag: str) -> list[str]:
    """Read domain names from the best run's step5_domains.json."""
    data = read_domains_json(dataset_dir, run_tag)
    if data is None:
        return []
    return [
        str(d.get("domain_name") or d.get("name") or "").strip()
        for d in data if isinstance(d, dict)
    ]


def count_single_table_domains(dataset_dir: Path, run_tag: str) -> int | None:
    """Count domains in a run whose tables list contains exactly one table."""
    data = read_domains_json(dataset_dir, run_tag)
    if data is None:
        return None
    count = 0
    for domain in data:
        if not isinstance(domain, dict):
            continue
        tables = domain.get("tables")
        if isinstance(tables, list) and len(tables) == 1:
            count += 1
    return count


def print_dataset_valid_configs(record: dict) -> None:
    """Print this dataset's Q > 0.3 configurations ranked by tables/domain."""
    configs = record.get("valid_configs", [])
    print("         Q > 0.3 configs ranked by Tables/Domain:")
    print("         Rank  theta_A  theta_T  resolution  run_tag                 status   Q       Tables  Domains  Single-table Domains  Tables/Domain")

    if not configs:
        print("           none")
        return

    for rank, cfg in enumerate(configs, 1):
        ratio = cfg.get("table_domain_ratio")
        ratio_text = f"{ratio:.4f}" if ratio is not None else "n/a"
        q_text = f"{cfg['Q']:.4f}" if cfg.get("Q") is not None else "n/a"
        single_table_domains = cfg.get("single_table_domains")
        single_text = str(single_table_domains) if single_table_domains is not None else "n/a"
        print(
            f"         {rank:>4}  "
            f"{cfg['theta_a']!s:>7}  "
            f"{cfg['theta_t']!s:>7}  "
            f"{cfg['resolution']!s:>10}  "
            f"{str(cfg['run_tag']):<22}  "
            f"{str(cfg['status']):<7}  "
            f"{q_text:>6}  "
            f"{cfg['n_tables']!s:>6}  "
            f"{cfg['n_domains']!s:>7}  "
            f"{single_text:>20}  "
            f"{ratio_text:>13}"
        )

# =============================================================================
# Build summary
# =============================================================================

def build_summary(root: Path, datasets: list[str]) -> list[dict]:
    records = []
    for name in datasets:
        dataset_dir = root / name
        if not dataset_dir.is_dir():
            continue

        best = read_best_from_xlsx(dataset_dir)
        if best is None:
            print(f"  [SKIP] {name:30s} — tune_params_results.xlsx not found or empty")
            continue

        run_tag = str(best.get("run_tag", ""))
        domain_names = read_domain_names(dataset_dir, run_tag) if run_tag else []
        stats = read_dataset_stats(dataset_dir)

        record = {"Dataset": name}
        record.update(best)
        record["domain_names"] = domain_names
        record["valid_configs"] = read_valid_configs_from_xlsx(dataset_dir)
        record.update(stats)
        record["table_count_check"] = table_count_check(
            record.get("domain_table_sum"),
            record.get("total_tables"),
            record.get("csv_tables"),
        )

        print(f"  [OK]   {name:30s}  Q={best['Q']:.4f}  "
              f"domains={best['n_domains']}  tag={run_tag}"
              f"  tables={stats['total_tables']}  cols={stats['total_columns']}")
        print(
            "         Table count check: "
            f"domain_sum={record.get('domain_table_sum')}  "
            f"schema_tables={record.get('total_tables')}  "
            f"csv_tables={record.get('csv_tables')}  "
            f"=> {record['table_count_check']}"
        )
        print_dataset_valid_configs(record)
        print()

        records.append(record)
    return records

# =============================================================================
# Write Excel
# =============================================================================

def write_excel(records: list[dict], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Best Configurations"

    dark_blue  = "1F3864"
    mid_blue   = "2F5496"
    green_fill = PatternFill("solid", fgColor="C6EFCE")
    good_fill  = PatternFill("solid", fgColor="EBF3E8")
    alt_fill   = PatternFill("solid", fgColor="EEF3FA")
    white_fill = PatternFill("solid", fgColor="FFFFFF")
    thin       = Side(style="thin", color="BFBFBF")
    border     = Border(left=thin, right=thin, top=thin, bottom=thin)
    center     = Alignment(horizontal="center", vertical="center")
    left       = Alignment(horizontal="left",   vertical="center")

    # Title
    ws.merge_cells("A1:S1")
    t = ws["A1"]
    t.value     = (f"DomainMiner - Best Configuration per Dataset   |   "
                   f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    t.font      = Font(name="Arial", bold=True, size=12, color="FFFFFF")
    t.fill      = PatternFill("solid", fgColor=dark_blue)
    t.alignment = center
    ws.row_dimensions[1].height = 22

    # Headers
    headers = [
        ("Dataset",            18), ("theta_A",       9), ("theta_T",     9),
        ("Resolution",         11), ("Run Tag",       24),
        ("Total Tables",       12), ("Total Columns", 13),
        ("CSV Tables",         10), ("Domain Table Sum", 16),
        ("Table Count Check",  17),
        ("Tables G_T",          9), ("Edges G_T",     10), ("Domains",     9),
        ("Single-table Domains", 21), ("Tables/Domain", 14), ("Q",         9),
        ("Status",             10), ("Time (s)",     10),
        ("Discovered Domains", 70),
    ]
    for col, (label, width) in enumerate(headers, 1):
        c = ws.cell(row=2, column=col, value=label)
        c.font      = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        c.fill      = PatternFill("solid", fgColor=mid_blue)
        c.alignment = center
        c.border    = border
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[2].height = 16

    # Data rows sorted by Q descending
    records_sorted = sorted(records, key=lambda r: r.get("Q") or 0, reverse=True)
    best_q = records_sorted[0]["Q"] if records_sorted else None

    for i, rec in enumerate(records_sorted):
        row    = i + 3
        q      = rec.get("Q")
        is_best = q == best_q
        fill   = (green_fill if is_best else
                  good_fill  if (q and q >= 0.3) else
                  alt_fill   if i % 2 == 0 else white_fill)

        def dc(col, value, fmt=None, bold=False, wrap=False):
            c = ws.cell(row=row, column=col, value=value)
            c.font      = Font(name="Arial", size=10 if col < 16 else 9, bold=bold)
            c.fill      = fill
            c.border    = border
            c.alignment = (Alignment(horizontal="left", vertical="center",
                                     wrap_text=wrap)
                           if col in (1, 19) else center)
            if fmt:
                c.number_format = fmt

        dc(1,  rec["Dataset"],    bold=is_best)
        dc(2,  rec["theta_a"])
        dc(3,  rec["theta_t"])
        dc(4,  rec["resolution"])
        dc(5,  rec["run_tag"])
        dc(6,  rec.get("total_tables"))
        dc(7,  rec.get("total_columns"))
        dc(8,  rec.get("csv_tables"))
        dc(9,  rec.get("domain_table_sum"))
        dc(10, rec.get("table_count_check"), bold=rec.get("table_count_check") == "MISMATCH")
        dc(11, rec["n_tables"])
        dc(12, rec["n_edges"])
        dc(13, rec["n_domains"])
        dc(14, rec.get("single_table_domains"))
        dc(15, round(rec["table_domain_ratio"], 4)
           if rec.get("table_domain_ratio") is not None else None,
           fmt="0.0000")
        dc(16, round(q, 4) if q is not None else None,
           fmt="0.0000", bold=is_best)
        dc(17, rec["status"])
        dc(18, rec["elapsed_s"])

        names      = rec.get("domain_names", [])
        domain_str = "  |  ".join(f"D{i}: {n}" for i, n in enumerate(names))
        dc(19, domain_str, wrap=True)

        ws.row_dimensions[row].height = max(15, 14 * ((len(domain_str) // 70) + 1))

    # Legend
    lr = len(records_sorted) + 4
    ws.merge_cells(f"A{lr}:S{lr}")
    leg = ws[f"A{lr}"]
    leg.value = ("Colour scale:  Green = highest Q overall   |   "
                 "Light green = Q ≥ 0.3 (valid)   |   Blue = Q < 0.3")
    leg.font      = Font(name="Arial", italic=True, size=9, color="595959")
    leg.alignment = left

    ws.freeze_panes = "A3"
    write_valid_configs_sheet(wb, records, dark_blue, mid_blue, border, center, left)
    wb.save(output_path)
    print(f"\n  Saved: {output_path}")


def write_valid_configs_sheet(
    wb: openpyxl.Workbook,
    records: list[dict],
    dark_blue: str,
    mid_blue: str,
    border: Border,
    center: Alignment,
    left: Alignment,
) -> None:
    """Add all Q > 0.3 configurations ranked by tables/domain per dataset."""
    ws = wb.create_sheet("Q_gt_0.3 by Ratio")

    ws.merge_cells("A1:P1")
    t = ws["A1"]
    t.value = "Configurations with Q > 0.3, ranked by Tables/Domain ratio within each dataset; includes table-count checks"
    t.font = Font(name="Arial", bold=True, size=12, color="FFFFFF")
    t.fill = PatternFill("solid", fgColor=dark_blue)
    t.alignment = center
    ws.row_dimensions[1].height = 22

    headers = [
        ("Dataset", 22),
        ("Rank", 8),
        ("theta_A", 9),
        ("theta_T", 9),
        ("Resolution", 11),
        ("Run Tag", 24),
        ("Status", 10),
        ("Table Count Check", 17),
        ("Domain Table Sum", 16),
        ("Schema Tables", 13),
        ("CSV Tables", 10),
        ("Q", 9),
        ("Tables", 10),
        ("Domains", 10),
        ("Single-table Domains", 21),
        ("Tables/Domain", 14),
    ]
    for col, (label, width) in enumerate(headers, 1):
        c = ws.cell(row=2, column=col, value=label)
        c.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor=mid_blue)
        c.alignment = center
        c.border = border
        ws.column_dimensions[get_column_letter(col)].width = width

    row = 3
    records_by_dataset = sorted(records, key=lambda r: r["Dataset"].lower())
    for rec in records_by_dataset:
        configs = rec.get("valid_configs", [])
        if not configs:
            values = [
                rec["Dataset"],
                "none",
                None,
                None,
                None,
                None,
                None,
                rec.get("table_count_check"),
                rec.get("domain_table_sum"),
                rec.get("total_tables"),
                rec.get("csv_tables"),
                None,
                None,
                None,
                None,
                None,
            ]
            for col, value in enumerate(values, 1):
                c = ws.cell(row=row, column=col, value=value)
                c.font = Font(name="Arial", size=10, italic=(col == 2))
                c.border = border
                c.alignment = left if col in (1, 6, 7, 8) else center
            row += 2
            continue

        for rank, cfg in enumerate(configs, 1):
            values = [
                rec["Dataset"],
                rank,
                cfg["theta_a"],
                cfg["theta_t"],
                cfg["resolution"],
                cfg["run_tag"],
                cfg["status"],
                rec.get("table_count_check"),
                rec.get("domain_table_sum"),
                rec.get("total_tables"),
                rec.get("csv_tables"),
                round(cfg["Q"], 4) if cfg["Q"] is not None else None,
                cfg["n_tables"],
                cfg["n_domains"],
                cfg.get("single_table_domains"),
                round(cfg["table_domain_ratio"], 4)
                if cfg["table_domain_ratio"] is not None else None,
            ]
            for col, value in enumerate(values, 1):
                c = ws.cell(row=row, column=col, value=value)
                c.font = Font(name="Arial", size=10)
                c.border = border
                c.alignment = left if col in (1, 6, 7, 8) else center
                if col in (12, 16):
                    c.number_format = "0.0000"
            row += 1

        if configs:
            row += 1

    ws.freeze_panes = "A3"


def resolve_output_path(output_path: Path) -> Path:
    """Keep relative Excel outputs in the project results folder."""
    if output_path.is_absolute():
        return output_path

    if output_path.parts and output_path.parts[0].lower() == "results":
        return PROJECT_ROOT / output_path

    return PROJECT_ROOT / "results" / output_path


# =============================================================================
# Main
# =============================================================================

def main() -> None:
    parser = argparse.ArgumentParser(
        description="List best configuration per dataset with domain names."
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=PROJECT_ROOT / "results" / "best_configs_summary.xlsx",
        help="Excel output path (default: results/best_configs_summary.xlsx).",
    )
    parser.add_argument("--datasets", nargs="+", default=None)
    parser.add_argument(
        "--root",
        type=Path,
        default=DATALAKES_ROOT,
        help="Dataset root containing dataset folders (default: Datalakes/).",
    )
    args = parser.parse_args()

    root     = args.root.resolve()
    output   = resolve_output_path(args.output)
    datasets = args.datasets or discover_datasets(root)

    print(f"\nDomainMiner - Best Configuration per Dataset")
    print(f"Root   : {root}")
    print(f"Output : {output}")
    print(f"Found  : {len(datasets)} candidate dataset folder(s)\n")

    records = build_summary(root, datasets)
    if not records:
        print("\n[ERROR] No results found.")
        return

    print(f"\n  {len(records)} dataset(s) found.\n")
    write_excel(records, output)


if __name__ == "__main__":
    main()
