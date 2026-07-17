"""
list_derived_weights.py
=======================
Collects derived_weights.csv from every dataset folder and writes a
summary Excel file to the results folder.

For each dataset it reads:
  <Dataset>/ccm_output/derived_weights.csv

which contains the variance-based weights w1 (P_stat), w2 (P_name),
w3 (P_sem) used by the CCM pipeline.

OUTPUT
------
  derived_weights_summary.xlsx   — one row per dataset, written to the
                                   project root (same folder as this script)

USAGE
-----
  python tools/list_derived_weights.py

  # Specify a custom output path
  python tools/list_derived_weights.py --output results/weights.xlsx

  # Include only specific datasets
  python tools/list_derived_weights.py --datasets Sakila Northwind Chinook
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# =============================================================================
# Configuration
# =============================================================================

PROJECT_ROOT = Path(__file__).resolve().parents[1]
PIPELINE_DIR = PROJECT_ROOT / "pipeline"
if str(PIPELINE_DIR) not in sys.path:
    sys.path.insert(0, str(PIPELINE_DIR))

from path_utils import DATALAKES_ROOT

# Known dataset folders (all project folders in root)
KNOWN_DATASETS = [
    "Sakila",
    "Northwind",
    "Chinook",
    "DellStore2",
    "adventure_works",
    "WideWorldImporters",
    "Employees",
    "TPCDS",
    "Airportdb",
    "FDA_AdverseEvents",
    "StackOverflowDataDump",
    "eicu",
    "mimiciv",
    "Synthea",
    "tcph",
]

WEIGHTS_FILE = Path("ccm_output") / "derived_weights.csv"

# =============================================================================
# Load weights for one dataset
# =============================================================================

def load_weights(dataset_dir: Path) -> dict | None:
    """
    Read derived_weights.csv for a dataset.
    Returns a dict with w1, w2, w3 and variance columns, or None if missing.
    """
    weights_path = dataset_dir / WEIGHTS_FILE
    if not weights_path.exists():
        return None

    try:
        df = pd.read_csv(weights_path)

        # Normalise column names to lowercase for robustness
        df.columns = df.columns.str.strip().str.lower()

        row = {}

        # Extract w1, w2, w3 — handle both 'w1'/'w2'/'w3' and 'weight' column layouts
        if "w1" in df.columns:
            row["w1_stat"]  = float(df["w1"].iloc[0])
            row["w2_name"]  = float(df["w2"].iloc[0])
            row["w3_sem"]   = float(df["w3"].iloc[0])
        elif "weight" in df.columns and "measure" in df.columns:
            # Long format: measure, weight
            df_idx = df.set_index("measure")["weight"]
            row["w1_stat"] = float(df_idx.get("w1", df_idx.get("stat", float("nan"))))
            row["w2_name"] = float(df_idx.get("w2", df_idx.get("name", float("nan"))))
            row["w3_sem"]  = float(df_idx.get("w3", df_idx.get("sem",  float("nan"))))
        else:
            # Try to read first three numeric columns as w1, w2, w3
            nums = df.select_dtypes(include="number")
            if nums.shape[1] >= 3:
                row["w1_stat"] = float(nums.iloc[0, 0])
                row["w2_name"] = float(nums.iloc[0, 1])
                row["w3_sem"]  = float(nums.iloc[0, 2])
            else:
                return None

        # Optionally read variance columns if present
        for col, key in [("var_stat", "var_stat"), ("var_name", "var_name"), ("var_sem", "var_sem")]:
            if col in df.columns:
                row[key] = float(df[col].iloc[0])

        return row

    except Exception as e:
        print(f"  [WARN] Could not parse {weights_path}: {e}")
        return None


# =============================================================================
# Build summary dataframe
# =============================================================================

def build_summary(root: Path, datasets: list[str]) -> pd.DataFrame:
    records = []

    for name in datasets:
        dataset_dir = root / name
        if not dataset_dir.is_dir():
            continue

        weights = load_weights(dataset_dir)
        if weights is None:
            print(f"  [SKIP] {name:30s} — derived_weights.csv not found")
            continue

        record = {"Dataset": name}
        record.update(weights)

        # Dominant measure (highest weight)
        w_map = {
            "P_stat (w1)": weights.get("w1_stat", 0),
            "P_name (w2)": weights.get("w2_name", 0),
            "P_sem  (w3)": weights.get("w3_sem",  0),
        }
        record["Dominant"] = max(w_map, key=w_map.get)

        records.append(record)
        print(f"  [OK]   {name:30s}  w1={weights.get('w1_stat', '?'):.4f}  "
              f"w2={weights.get('w2_name', '?'):.4f}  w3={weights.get('w3_sem', '?'):.4f}")

    if not records:
        print("\n[ERROR] No derived_weights.csv found in any dataset folder.")
        sys.exit(1)

    df = pd.DataFrame(records)

    # Sort by w1 descending
    if "w1_stat" in df.columns:
        df = df.sort_values("w1_stat", ascending=False).reset_index(drop=True)

    return df


# =============================================================================
# Write Excel
# =============================================================================

def write_excel(df: pd.DataFrame, output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Derived Weights"

    # ── Styles ──────────────────────────────────────────────────────────────
    header_fill   = PatternFill("solid", fgColor="1F4E79")
    header_font   = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    subhdr_fill   = PatternFill("solid", fgColor="2E75B6")
    subhdr_font   = Font(name="Calibri", bold=True, color="FFFFFF", size=10)

    green_fill    = PatternFill("solid", fgColor="C6EFCE")
    yellow_fill   = PatternFill("solid", fgColor="FFEB9C")
    red_fill      = PatternFill("solid", fgColor="FFC7CE")

    center        = Alignment(horizontal="center", vertical="center")
    left          = Alignment(horizontal="left",   vertical="center")

    thin          = Side(style="thin", color="BFBFBF")
    border        = Border(left=thin, right=thin, top=thin, bottom=thin)

    body_font     = Font(name="Calibri", size=10)

    # ── Title row ────────────────────────────────────────────────────────────
    ws.merge_cells("A1:H1")
    title_cell = ws["A1"]
    title_cell.value = "DomainMiner - Derived Similarity Weights per Dataset"
    title_cell.font  = Font(name="Calibri", bold=True, size=13, color="1F4E79")
    title_cell.alignment = center
    ws.row_dimensions[1].height = 24

    # ── Column headers ────────────────────────────────────────────────────────
    has_var = "var_stat" in df.columns

    base_headers = [
        ("Dataset",         18),
        ("w1  P_stat",      12),
        ("w2  P_name",      12),
        ("w3  P_sem",       12),
        ("Dominant",        20),
    ]
    var_headers = [
        ("Var(P_stat)",     12),
        ("Var(P_name)",     12),
        ("Var(P_sem)",      12),
    ]
    headers = base_headers + (var_headers if has_var else [])

    for col_idx, (label, width) in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx, value=label)
        cell.fill      = subhdr_fill
        cell.font      = subhdr_font
        cell.alignment = center
        cell.border    = border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[2].height = 18

    # ── Data rows ─────────────────────────────────────────────────────────────
    col_order = ["Dataset", "w1_stat", "w2_name", "w3_sem", "Dominant"]
    if has_var:
        col_order += ["var_stat", "var_name", "var_sem"]

    for row_idx, row in df.iterrows():
        excel_row = row_idx + 3
        row_bg = PatternFill("solid", fgColor="EEF3FA" if row_idx % 2 == 0 else "FFFFFF")

        for col_idx, col_key in enumerate(col_order, start=1):
            val  = row.get(col_key, "")
            cell = ws.cell(row=excel_row, column=col_idx, value=val)
            cell.font      = body_font
            cell.border    = border
            cell.fill      = row_bg

            # Format floats
            if isinstance(val, float):
                cell.number_format = "0.0000"
                cell.alignment     = center

                # Colour-code weight columns (cols 2–4)
                if col_idx in (2, 3, 4):
                    if val >= 0.50:
                        cell.fill = green_fill
                    elif val >= 0.25:
                        cell.fill = yellow_fill
                    else:
                        cell.fill = red_fill
            else:
                cell.alignment = left if col_idx == 1 else center

        ws.row_dimensions[excel_row].height = 16

    # ── Legend ───────────────────────────────────────────────────────────────
    legend_row = len(df) + 4
    ws.merge_cells(f"A{legend_row}:H{legend_row}")
    legend_cell = ws[f"A{legend_row}"]
    legend_cell.value = (
        "Colour scale (weights):  "
        "Green ≥ 0.50 (dominant)   |   "
        "Yellow 0.25–0.49 (moderate)   |   "
        "Red < 0.25 (weak)"
    )
    legend_cell.font      = Font(name="Calibri", italic=True, size=9, color="595959")
    legend_cell.alignment = left

    note_row = legend_row + 1
    ws.merge_cells(f"A{note_row}:H{note_row}")
    note_cell = ws[f"A{note_row}"]
    note_cell.value = (
        "Weights are variance-based: w_k = Var(signal_k) / ΣVar.  "
        "w1=P_stat (statistical, z-score inverted),  "
        "w2=P_name (Levenshtein),  "
        "w3=P_sem (cosine on φ-matrix rows)."
    )
    note_cell.font      = Font(name="Calibri", italic=True, size=9, color="595959")
    note_cell.alignment = left

    # ── Freeze panes ─────────────────────────────────────────────────────────
    ws.freeze_panes = "A3"

    wb.save(output_path)
    print(f"\n  Saved: {output_path}")


# =============================================================================
# Main
# =============================================================================

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Collect derived_weights.csv from all dataset folders and write a summary Excel file."
    )
    parser.add_argument(
        "--output", type=Path, default=PROJECT_ROOT / "results" / "derived_weights_summary.xlsx",
        help="Output Excel file path (default: results/derived_weights_summary.xlsx)"
    )
    parser.add_argument(
        "--datasets", nargs="+", default=None,
        help="Specific dataset folder names to include (default: all known datasets)"
    )
    parser.add_argument(
        "--root", type=Path, default=DATALAKES_ROOT,
        help="Dataset root directory (default: Datalakes/)"
    )
    args = parser.parse_args()

    root     = args.root.resolve()
    datasets = args.datasets or KNOWN_DATASETS

    print(f"\nDomainMiner - Derived Weights Collector")
    print(f"Root   : {root}")
    print(f"Output : {args.output}\n")

    df = build_summary(root, datasets)

    print(f"\n  {len(df)} dataset(s) found with derived weights.\n")
    print(df.to_string(index=False))

    write_excel(df, args.output)


if __name__ == "__main__":
    main()
