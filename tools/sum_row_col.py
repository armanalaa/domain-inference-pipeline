"""
sum_row_col.py
--------------
Print total tables and total columns for one or more datasets.

Usage:
    python tools/sum_row_col.py
    python tools/sum_row_col.py Sakila Mondial

If no dataset is provided, every project folder containing CSV files is
summarized. The same concise result is written to an Excel file in results/.
"""

from __future__ import annotations

import argparse
import os
import sys
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


PROJECT_ROOT = Path(__file__).resolve().parents[1]
PIPELINE_DIR = PROJECT_ROOT / "pipeline"
if str(PIPELINE_DIR) not in sys.path:
    sys.path.insert(0, str(PIPELINE_DIR))

from path_utils import DATALAKES_ROOT, resolve_dataset_dir


def safe_name(name: str) -> str:
    return "".join(ch if ch.isalnum() or ch in ("-", "_") else "_" for ch in name)


def find_csv_files(root_dir: Path, recursive: bool) -> list[Path]:
    found = []
    if recursive:
        for dirpath, _dirnames, filenames in os.walk(root_dir):
            for fname in filenames:
                if fname.lower().endswith(".csv"):
                    found.append(Path(dirpath) / fname)
    else:
        try:
            for path in root_dir.iterdir():
                if path.is_file() and path.suffix.lower() == ".csv":
                    found.append(path)
        except Exception as exc:
            print(f"  Could not list directory {root_dir}: {exc}")
    return sorted(found)


def discover_datasets(subdir: str, recursive: bool) -> list[Path]:
    datasets = []
    for child in sorted(DATALAKES_ROOT.iterdir() if DATALAKES_ROOT.exists() else PROJECT_ROOT.iterdir()):
        if not child.is_dir():
            continue
        target_dir = child / subdir if subdir else child
        if target_dir.exists() and find_csv_files(target_dir, recursive):
            datasets.append(child)
    return datasets


def resolve_dataset(name_or_path: str) -> Path:
    return resolve_dataset_dir(name_or_path)


def count_csv_columns(csv_path: Path) -> int:
    """Return the number of columns in a CSV without reading its rows."""
    try:
        return len(pd.read_csv(csv_path, nrows=0).columns)
    except pd.errors.EmptyDataError:
        return 0
    except Exception as exc:
        print(f"  [WARN] Could not read {csv_path}: {exc}")
        return 0


def summarize_dataset(dataset_dir: Path, subdir: str, recursive: bool) -> dict:
    target_dir = dataset_dir / subdir if subdir else dataset_dir
    csv_files = find_csv_files(target_dir, recursive) if target_dir.exists() else []

    total_columns = 0
    for csv_path in csv_files:
        total_columns += count_csv_columns(csv_path)

    return {
        "Dataset": dataset_dir.name,
        "Total Tables": len(csv_files),
        "Total Columns": total_columns,
    }


def autosize_columns(writer: pd.ExcelWriter, sheet_name: str) -> None:
    ws = writer.book[sheet_name]
    for column_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 40)


def write_excel(output_path: Path, records: list[dict]) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    df = pd.DataFrame(records, columns=["Dataset", "Total Tables", "Total Columns"])
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Row Column Totals", index=False)
        ws = writer.book["Row Column Totals"]
        header_fill = PatternFill("solid", fgColor="1F4E79")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        ws.freeze_panes = "A2"
        autosize_columns(writer, "Row Column Totals")


def print_summary(records: list[dict]) -> None:
    print(f"{'Dataset':30s} {'Total Tables':>15s} {'Total Columns':>15s}")
    print("-" * 64)
    for record in records:
        print(
            f"{record['Dataset']:30s} "
            f"{record['Total Tables']:15,} "
            f"{record['Total Columns']:15,}"
        )


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Print total tables and total columns for dataset CSV tables."
    )
    parser.add_argument(
        "datasets",
        nargs="*",
        help="Dataset folder names or full paths. If omitted, all datasets with CSV files are included.",
    )
    parser.add_argument(
        "--subdir",
        default="csv",
        help='Subfolder to look in, relative to each dataset folder (default: "csv").',
    )
    parser.add_argument(
        "--recursive",
        action="store_true",
        help="Search CSV subfolders recursively.",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=None,
        help="Excel output path (default: results/row_col_totals.xlsx).",
    )
    args = parser.parse_args()

    dataset_dirs = (
        [resolve_dataset(name) for name in args.datasets]
        if args.datasets
        else discover_datasets(args.subdir, args.recursive)
    )

    if not dataset_dirs:
        print("No dataset folders with CSV files found.")
        sys.exit(0)

    records = []
    for dataset_dir in dataset_dirs:
        if not dataset_dir.exists():
            print(f"  [SKIP] Dataset folder not found: {dataset_dir}")
            continue
        records.append(summarize_dataset(dataset_dir, args.subdir, args.recursive))

    if not records:
        print("No datasets summarized.")
        sys.exit(0)

    print_summary(records)

    output_path = args.output or PROJECT_ROOT / "results" / "row_col_totals.xlsx"
    write_excel(output_path, records)
    print(f"\nExcel report saved to: {output_path}")


if __name__ == "__main__":
    main()
