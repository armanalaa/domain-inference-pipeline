"""
build_tune_params_results.py
============================
Reads all completed tA*_tT*_r* subfolders in ccm_output/ and writes:
  - ccm_output/tune_params_results.xlsx
  - ccm_output/tune_params_summary.txt

Run from INSIDE the dataset folder:
    cd Synthea
    python ..\build_tune_params_results.py

Or pass the dataset folder explicitly:
    python build_tune_params_results.py --dataset_dir Synthea
"""

from __future__ import annotations

import argparse
import re
import sys
from datetime import datetime
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parent.parent
PIPELINE_DIR = PROJECT_ROOT / "pipeline"
if str(PIPELINE_DIR) not in sys.path:
    sys.path.insert(0, str(PIPELINE_DIR))

from path_utils import dataset_display_name, resolve_dataset_dir


# =============================================================================
# Parse step5_report.txt
# =============================================================================

def parse_report(report_path: Path) -> dict:
    result = {"Q": None, "n_domains": None, "n_tables": None, "n_edges": None}
    if not report_path.exists():
        return result
    text = report_path.read_text(encoding="utf-8")
    for line in text.splitlines():
        m = re.search(r"Modularity Q\s*:\s*(-?[\d.]+)", line)
        if m:
            result["Q"] = float(m.group(1))
        m = re.search(r"Domains discovered\s*:\s*(\d+)", line)
        if m:
            result["n_domains"] = int(m.group(1))
        m = re.search(r"Tables in G_T\s*:\s*(\d+)", line)
        if m:
            result["n_tables"] = int(m.group(1))
        m = re.search(r"Edges in G_T\s*:\s*(\d+)", line)
        if m:
            result["n_edges"] = int(m.group(1))
    return result


def make_run_tag(theta_a: float, theta_t: float, resolution: float) -> str:
    return f"tA{theta_a}_tT{theta_t}_r{resolution}"


# =============================================================================
# Scan ccm_output/ for completed runs
# =============================================================================

def scan_runs(dataset_dir: Path) -> list[dict]:
    out_dir = dataset_dir / "ccm_output"
    if not out_dir.exists():
        print(f"[ERROR] ccm_output/ not found in {dataset_dir.resolve()}")
        return []

    results = []
    # Match folders like tA0.65_tT0.70_r1.2
    pattern = re.compile(r"tA([\d.]+)_tT([\d.]+)_r([\d.]+)")

    for folder in sorted(out_dir.iterdir()):
        if not folder.is_dir():
            continue
        m = pattern.fullmatch(folder.name)
        if not m:
            continue

        theta_a    = float(m.group(1))
        theta_t    = float(m.group(2))
        resolution = float(m.group(3))
        tag        = folder.name

        report = folder / "step5_report.txt"
        parsed = parse_report(report)

        if parsed["Q"] is not None:
            status = "ok"
        elif report.exists():
            status = "FAILED"   # report exists but Q not found
        else:
            status = "FAILED"   # no report at all

        results.append({
            "theta_a":    theta_a,
            "theta_t":    theta_t,
            "resolution": resolution,
            "run_tag":    tag,
            "status":     status,
            "elapsed_s":  None,   # not available when rebuilding
            "Q":          parsed["Q"],
            "n_domains":  parsed["n_domains"],
            "n_tables":   parsed["n_tables"],
            "n_edges":    parsed["n_edges"],
        })

    return results


# =============================================================================
# Write Excel
# =============================================================================

def write_excel(results: list[dict], path: Path, dataset_dir: str) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("[WARNING] openpyxl not installed - skipping Excel output.")
        print("          Run: pip install openpyxl")
        return

    completed        = [r for r in results if r["Q"] is not None]
    failed           = [r for r in results if r["status"] == "FAILED"]
    completed_sorted = sorted(completed, key=lambda r: r["Q"], reverse=True)
    best             = completed_sorted[0] if completed_sorted else None

    C_HEADER = "2F5496"
    C_BEST   = "E2EFDA"
    C_GOOD   = "EBF3E8"
    C_FAILED = "FCE4D6"
    C_TITLE  = "D9E1F2"

    thin   = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def header_cell(ws, row, col, value, width=None):
        c = ws.cell(row=row, column=col, value=value)
        c.font      = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        c.fill      = PatternFill("solid", start_color=C_HEADER)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = border
        if width:
            ws.column_dimensions[get_column_letter(col)].width = width
        return c

    def data_cell(ws, row, col, value, fill_color=None,
                  bold=False, number_format=None, align="center"):
        c = ws.cell(row=row, column=col, value=value)
        c.font      = Font(name="Arial", bold=bold, size=10)
        c.alignment = Alignment(horizontal=align, vertical="center")
        c.border    = border
        if fill_color:
            c.fill = PatternFill("solid", start_color=fill_color)
        if number_format:
            c.number_format = number_format
        return c

    wb = Workbook()

    # ── Sheet 1: Results ──────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Results"
    ws1.row_dimensions[1].height = 22

    ws1.merge_cells("A1:J1")
    title = ws1["A1"]
    title.value     = (f"CCM Parameter Tuning - {dataset_dir}   |   "
                       f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    title.font      = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    title.fill      = PatternFill("solid", start_color="1F3864")
    title.alignment = Alignment(horizontal="center", vertical="center")

    headers = ["theta_A", "theta_T", "Resolution", "Run Tag",
               "Tables", "Edges G_T", "Domains", "Q", "Status", "Time (s)"]
    widths  = [9, 9, 11, 28, 8, 10, 9, 9, 10, 10]
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        header_cell(ws1, 2, col, h, width=w)

    all_sorted = completed_sorted + failed
    for i, r in enumerate(all_sorted, 3):
        is_best   = best and r is best
        is_good   = r["Q"] is not None and r["Q"] > 0.3 and not is_best
        is_failed = r["status"] == "FAILED"
        fill = (C_BEST   if is_best   else
                C_GOOD   if is_good   else
                C_FAILED if is_failed else None)

        data_cell(ws1, i, 1,  r["theta_a"],    fill)
        data_cell(ws1, i, 2,  r["theta_t"],    fill)
        data_cell(ws1, i, 3,  r["resolution"], fill)
        data_cell(ws1, i, 4,  r["run_tag"],    fill, align="left")
        data_cell(ws1, i, 5,  r["n_tables"],   fill)
        data_cell(ws1, i, 6,  r["n_edges"],    fill)
        data_cell(ws1, i, 7,  r["n_domains"],  fill)
        data_cell(ws1, i, 8,
                  round(r["Q"], 4) if r["Q"] is not None else None,
                  fill, bold=is_best, number_format="0.0000")
        data_cell(ws1, i, 9,  r["status"],     fill)
        data_cell(ws1, i, 10, r["elapsed_s"],  fill)

    ws1.freeze_panes = "A3"

    # ── Sheet 2: Summary ──────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")

    def title_row(ws, row, text):
        ws.merge_cells(start_row=row, start_column=1,
                       end_row=row,   end_column=6)
        c = ws.cell(row=row, column=1, value=text)
        c.font      = Font(name="Arial", bold=True, size=11, color="FFFFFF")
        c.fill      = PatternFill("solid", start_color="1F3864")
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 20

    def kv(ws, row, key, value):
        k = ws.cell(row=row, column=1, value=key)
        k.font      = Font(name="Arial", bold=True, size=10)
        k.fill      = PatternFill("solid", start_color=C_TITLE)
        k.alignment = Alignment(horizontal="left")
        k.border    = border
        v = ws.cell(row=row, column=2, value=value)
        v.font      = Font(name="Arial", size=10)
        v.alignment = Alignment(horizontal="left")
        v.border    = border
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 30

    row = 1
    title_row(ws2, row, f"Parameter Tuning Summary - {dataset_dir}")
    row += 1
    kv(ws2, row, "Generated",  datetime.now().strftime("%Y-%m-%d %H:%M:%S")); row += 1
    kv(ws2, row, "Total runs", len(results));    row += 1
    kv(ws2, row, "Completed",  len(completed));  row += 1
    kv(ws2, row, "Failed",     len(failed));     row += 1
    kv(ws2, row, "Q > 0.3",    sum(1 for r in completed if r["Q"] > 0.3)); row += 2

    if best:
        title_row(ws2, row, "Best Run"); row += 1
        kv(ws2, row, "Run tag",    best["run_tag"]);    row += 1
        kv(ws2, row, "Q",          round(best["Q"], 4)); row += 1
        kv(ws2, row, "Domains",    best["n_domains"]);  row += 1
        kv(ws2, row, "Tables",     best["n_tables"]);   row += 1
        kv(ws2, row, "Edges G_T",  best["n_edges"]);    row += 1
        kv(ws2, row, "theta_A",    best["theta_a"]);    row += 1
        kv(ws2, row, "theta_T",    best["theta_t"]);    row += 1
        kv(ws2, row, "Resolution", best["resolution"]); row += 2

    if failed:
        title_row(ws2, row, "Failed Runs"); row += 1
        for r in failed:
            kv(ws2, row, r["run_tag"], "FAILED"); row += 1

    wb.save(path)
    print(f"Excel report  -> {path}")


# =============================================================================
# Write plain-text summary
# =============================================================================

def write_summary(results: list[dict], path: Path, dataset_dir: str) -> None:
    completed = [r for r in results if r["Q"] is not None]
    failed    = [r for r in results if r["status"] == "FAILED"]

    if not completed:
        path.write_text("No completed runs to summarise.\n", encoding="utf-8")
        print(f"Summary       -> {path}  (no completed runs)")
        return

    completed.sort(key=lambda r: r["Q"], reverse=True)
    best = completed[0]

    lines = [
        "=" * 74,
        f"CCM Parameter Tuning Summary - {dataset_dir}",
        f"Generated : {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        f"Runs: {len(results)}   Completed: {len(completed)}   Failed: {len(failed)}",
        "=" * 74,
        "",
        f"  {'theta_A':>7}  {'theta_T':>7}  {'Res':>5}  {'Domains':>7}  "
        f"{'Edges':>6}  {'Q':>7}  Flag",
        "  " + "-" * 62,
    ]

    for r in completed:
        flag = "  < BEST" if r is best else ("  OK Q>0.3" if r["Q"] > 0.3 else "")
        lines.append(
            f"  {r['theta_a']:>7}  {r['theta_t']:>7}  {r['resolution']:>5}  "
            f"{str(r['n_domains']):>7}  {str(r['n_edges']):>6}  "
            f"{r['Q']:>7.4f}{flag}"
        )

    if failed:
        lines += ["", "  Failed runs:"]
        for r in failed:
            lines.append(f"    {r['run_tag']}")

    lines += ["", "=" * 74]
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    print(f"Text summary  -> {path}")


# =============================================================================
# Entry point
# =============================================================================

if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Build tune_params_results.xlsx and tune_params_summary.txt "
                    "from existing ccm_output/ run folders."
    )
    parser.add_argument(
        "--dataset_dir",
        default=".",
        help="Dataset folder to scan (default: current directory '.')",
    )
    args = parser.parse_args()

    dataset_path = resolve_dataset_dir(args.dataset_dir)
    dataset_name = dataset_display_name(args.dataset_dir, dataset_path)

    print(f"\nScanning: {dataset_path / 'ccm_output'}")

    results = scan_runs(dataset_path)

    if not results:
        print("No tA*_tT*_r* run folders found in ccm_output/. Nothing to do.")
        raise SystemExit(1)

    print(f"Found {len(results)} run(s): "
          f"{sum(1 for r in results if r['Q'] is not None)} completed, "
          f"{sum(1 for r in results if r['status'] == 'FAILED')} failed\n")

    out_dir = dataset_path / "ccm_output"
    write_excel(results,  out_dir / "tune_params_results.xlsx", dataset_name)
    write_summary(results, out_dir / "tune_params_summary.txt", dataset_name)

    print("\nDone.")
