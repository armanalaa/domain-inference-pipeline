"""
tune_params.py
==============
Runs run_pipeline.py across a grid of theta_A, theta_T, and resolution values,
collecting Q and n_domains from each run's step5_report.txt.

Produces:
  tune_params_results.xlsx  — formatted Excel report (Results + Summary sheets)
  tune_params_summary.txt   — plain-text summary, best run highlighted

Usage (from validation/ folder):
  # Default grid — edit THETA_A, THETA_T, RESOLUTION lists below
  python tune_params.py --dataset_dir adventure_works

  # Override grid from command line
  python tune_params.py --dataset_dir adventure_works ^
      --theta_a 0.65 ^
      --theta_t 0.60 0.65 0.70 0.75 ^
      --resolution 1.0 1.2 1.5 2.0

  # Skip LLM (fast Q-only search, no domain labels)
  python tune_params.py --dataset_dir adventure_works --no_llm

  # Dry run — print commands without executing
  python tune_params.py --dataset_dir adventure_works --dry_run
"""

from __future__ import annotations

import argparse
import re
import subprocess
import sys
import time
from datetime import datetime
from itertools import product
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parent.parent
PIPELINE_DIR = PROJECT_ROOT / "pipeline"
if str(PIPELINE_DIR) not in sys.path:
    sys.path.insert(0, str(PIPELINE_DIR))

from path_utils import dataset_display_name, resolve_dataset_dir

# =============================================================================
# DEFAULT GRID — edit these lists to change the search space
# =============================================================================
DEFAULT_THETA_A    = [0.65]
DEFAULT_THETA_T    = [0.60, 0.65, 0.70, 0.75]
DEFAULT_RESOLUTION = [1.0, 1.2, 1.5, 2.0]

# =============================================================================
# Parse step5_report.txt for Q and n_domains
# =============================================================================

def parse_report(report_path: Path) -> dict:
    """Extract Q and n_domains from step5_report.txt."""
    result = {"Q": None, "n_domains": None, "n_tables": None, "n_edges": None}
    if not report_path.exists():
        return result
    text = report_path.read_text(encoding="utf-8")
    for line in text.splitlines():
        m = re.search(r"Modularity Q\s*:\s*([\d.]+)", line)
        if m:
            result["Q"] = float(m.group(1))
        m = re.search(r"Domains discovered\s*:\s*(\d+)", line)
        if m:
            result["n_domains"] = int(m.group(1))
        m = re.search(r"Tables in G_T\s*:\s*(\d+)", line)
        if m:
            result["n_tables"] = int(m.group(1))
        m = re.search(r"Edges in G_T\s*:\s*(\d+)", line)
        if m:
            result["n_edges"] = int(m.group(1))
    return result


# =============================================================================
# Build run tag (mirrors pipeline_utils.make_run_tag)
# =============================================================================

def make_run_tag(theta_a: float, theta_t: float, resolution: float) -> str:
    return f"tA{theta_a}_tT{theta_t}_r{resolution}"


# =============================================================================
# Completion check
# =============================================================================

# All five files must exist and be non-empty for a run to be considered complete
_COMPLETION_FILES = [
    "step5_report.txt",
    "step5_domains.json",
    "step5_table_domain.csv",
    "step5_column_domain.csv",
    "step4_graph_edges.csv",
]

def is_run_complete(script_dir: Path, dataset_dir: str,
                    out_dir: str, tag: str) -> bool:
    """Return True only if all expected Step 5 outputs exist and are non-empty."""
    run_dir = script_dir / dataset_dir / out_dir / tag
    if not run_dir.exists():
        return False
    for fname in _COMPLETION_FILES:
        p = run_dir / fname
        if not p.exists() or p.stat().st_size == 0:
            return False
    # Also verify step5_report.txt contains a Q value (not truncated)
    report = run_dir / "step5_report.txt"
    try:
        text = report.read_text(encoding="utf-8")
        if "Modularity Q" not in text:
            return False
    except Exception:
        return False
    return True

def run_one(
    script_dir:   Path,
    dataset_dir:  str,
    schema:       str,
    knowledge:    str | None,
    theta_a:      float,
    theta_t:      float,
    resolution:   float,
    model:        str,
    ollama_url:   str,
    no_llm:       bool,
    out_dir:      str,
    dry_run:      bool,
    start_from:   str | None,
    force:        bool = False,
) -> dict:
    """Run pipeline for one parameter combination. Returns result dict."""

    tag = make_run_tag(theta_a, theta_t, resolution)

    # ── Checkpoint: skip if already fully completed ───────────────────────────
    if not force and not dry_run and is_run_complete(script_dir, dataset_dir, out_dir, tag):
        report_path = script_dir / dataset_dir / out_dir / tag / "step5_report.txt"
        parsed = parse_report(report_path)
        print()
        print(f"  SKIP (already complete): {tag}")
        print(f"       Q={parsed['Q']}  domains={parsed['n_domains']}")
        return {
            "theta_a":    theta_a,
            "theta_t":    theta_t,
            "resolution": resolution,
            "run_tag":    tag,
            "Q":          parsed["Q"],
            "n_domains":  parsed["n_domains"],
            "n_tables":   parsed["n_tables"],
            "n_edges":    parsed["n_edges"],
            "status":     "skipped",
            "elapsed_s":  0,
        }

    cmd = [
        sys.executable, str(script_dir / "scripts" / "run_pipeline.py"),
        "--dataset_dir", dataset_dir,
        "--schema",      schema,
        "--theta_a",     str(theta_a),
        "--theta_t",     str(theta_t),
        "--resolution",  str(resolution),
        "--model",       model,
        "--ollama_url",  ollama_url,
        "--out_dir",     out_dir,
        "--clean",
    ]
    if knowledge:
        cmd += ["--knowledge", knowledge]
    if no_llm:
        cmd.append("--no_llm")
    if start_from:
        cmd += ["--start_from", start_from]

    print()
    print("=" * 70)
    print(f"  RUN: {tag}")
    print(f"  CMD: {' '.join(cmd)}")
    print("=" * 70)

    result = {
        "theta_a":    theta_a,
        "theta_t":    theta_t,
        "resolution": resolution,
        "run_tag":    tag,
        "Q":          None,
        "n_domains":  None,
        "n_tables":   None,
        "n_edges":    None,
        "status":     "dry_run" if dry_run else "pending",
        "elapsed_s":  None,
    }

    if dry_run:
        return result

    t0 = time.time()
    proc = subprocess.run(cmd, text=True, cwd=str(script_dir))
    elapsed = round(time.time() - t0, 1)
    result["elapsed_s"] = elapsed

    if proc.returncode != 0:
        result["status"] = "FAILED"
        print(f"  FAILED after {elapsed}s")
        return result

    # Parse report from the run subfolder
    report_path = script_dir / dataset_dir / out_dir / tag / "step5_report.txt"
    parsed = parse_report(report_path)
    result.update(parsed)
    result["status"] = "ok"
    print(f"  OK  Q={result['Q']}  domains={result['n_domains']}  ({elapsed}s)")
    return result


# =============================================================================
# Write results
# =============================================================================

def write_excel(results: list[dict], path: Path, dataset_dir: str) -> None:
    """Write tuning results to a formatted Excel file with two sheets:
       - Results : one row per run, sorted by Q descending
       - Summary : best run highlighted, failed runs listed
    """
    from openpyxl import Workbook
    from openpyxl.styles import (Font, PatternFill, Alignment,
                                  Border, Side)
    from openpyxl.utils import get_column_letter

    completed = [r for r in results if r["Q"] is not None]
    failed    = [r for r in results if r["status"] == "FAILED"]
    completed_sorted = sorted(completed, key=lambda r: r["Q"], reverse=True)
    best = completed_sorted[0] if completed_sorted else None

    wb = Workbook()

    # ── colour palette ────────────────────────────────────────────────────────
    C_HEADER  = "2F5496"   # dark blue
    C_BEST    = "E2EFDA"   # light green
    C_GOOD    = "EBF3E8"   # very light green  (Q > 0.3)
    C_FAILED  = "FCE4D6"   # light orange
    C_TITLE   = "D9E1F2"   # light blue

    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def header_cell(ws, row, col, value, width=None):
        c = ws.cell(row=row, column=col, value=value)
        c.font      = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        c.fill      = PatternFill("solid", start_color=C_HEADER)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = border
        if width:
            ws.column_dimensions[get_column_letter(col)].width = width
        return c

    def data_cell(ws, row, col, value, fill_color=None,
                  bold=False, number_format=None, align="center"):
        c = ws.cell(row=row, column=col, value=value)
        c.font      = Font(name="Arial", bold=bold, size=10)
        c.alignment = Alignment(horizontal=align, vertical="center")
        c.border    = border
        if fill_color:
            c.fill = PatternFill("solid", start_color=fill_color)
        if number_format:
            c.number_format = number_format
        return c

    # =========================================================================
    # Sheet 1 — Results
    # =========================================================================
    ws1 = wb.active
    ws1.title = "Results"
    ws1.row_dimensions[1].height = 22

    # Title row
    ws1.merge_cells("A1:J1")
    title = ws1["A1"]
    title.value     = f"CCM Parameter Tuning - {dataset_dir}   |   Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    title.font      = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    title.fill      = PatternFill("solid", start_color="1F3864")
    title.alignment = Alignment(horizontal="center", vertical="center")

    # Column headers (row 2)
    headers = ["theta_A", "theta_T", "Resolution", "Run Tag",
               "Tables", "Edges G_T", "Domains", "Q",
               "Status", "Time (s)"]
    widths  = [9, 9, 11, 28, 8, 10, 9, 9, 10, 10]
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        header_cell(ws1, 2, col, h, width=w)

    # Data rows (row 3 onward) — sorted by Q descending
    all_sorted = completed_sorted + failed + [r for r in results
                                               if r["status"] not in ("ok","FAILED")]
    for i, r in enumerate(all_sorted, 3):
        is_best    = best and r is best
        is_good    = r["Q"] is not None and r["Q"] > 0.3 and not is_best
        is_failed  = r["status"] == "FAILED"
        is_skipped = r["status"] == "skipped"
        fill = (C_BEST   if is_best   else
                C_GOOD   if is_good   else
                "FFF2CC" if is_skipped else   # light yellow for skipped
                C_FAILED if is_failed  else None)

        data_cell(ws1, i, 1,  r["theta_a"],   fill)
        data_cell(ws1, i, 2,  r["theta_t"],   fill)
        data_cell(ws1, i, 3,  r["resolution"],fill)
        data_cell(ws1, i, 4,  r["run_tag"],   fill, align="left")
        data_cell(ws1, i, 5,  r["n_tables"],  fill)
        data_cell(ws1, i, 6,  r["n_edges"],   fill)
        data_cell(ws1, i, 7,  r["n_domains"], fill)
        q_cell = data_cell(ws1, i, 8,
                           round(r["Q"], 4) if r["Q"] is not None else None,
                           fill, bold=is_best, number_format="0.0000")
        data_cell(ws1, i, 9,  r["status"],    fill)
        data_cell(ws1, i, 10, r["elapsed_s"], fill)

    # Freeze panes below header
    ws1.freeze_panes = "A3"

    # =========================================================================
    # Sheet 2 — Summary
    # =========================================================================
    ws2 = wb.create_sheet("Summary")

    def title_row(ws, row, text):
        ws.merge_cells(start_row=row, start_column=1,
                       end_row=row,   end_column=6)
        c = ws.cell(row=row, column=1, value=text)
        c.font      = Font(name="Arial", bold=True, size=11, color="FFFFFF")
        c.fill      = PatternFill("solid", start_color="1F3864")
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 20

    def kv(ws, row, key, value):
        k = ws.cell(row=row, column=1, value=key)
        k.font      = Font(name="Arial", bold=True, size=10)
        k.fill      = PatternFill("solid", start_color=C_TITLE)
        k.alignment = Alignment(horizontal="left")
        k.border    = border
        v = ws.cell(row=row, column=2, value=value)
        v.font      = Font(name="Arial", size=10)
        v.alignment = Alignment(horizontal="left")
        v.border    = border
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 30

    row = 1
    title_row(ws2, row, f"Parameter Tuning Summary - {dataset_dir}")
    row += 1
    kv(ws2, row, "Generated",    datetime.now().strftime("%Y-%m-%d %H:%M:%S")); row += 1
    kv(ws2, row, "Total runs",   len(results));   row += 1
    kv(ws2, row, "Completed",    len(completed)); row += 1
    kv(ws2, row, "Failed",       len(failed));    row += 1
    kv(ws2, row, "Q > 0.3",      sum(1 for r in completed if r["Q"] > 0.3)); row += 2

    if best:
        title_row(ws2, row, "Best Run"); row += 1
        kv(ws2, row, "Run tag",    best["run_tag"]);    row += 1
        kv(ws2, row, "Q",          round(best["Q"],4)); row += 1
        kv(ws2, row, "Domains",    best["n_domains"]);  row += 1
        kv(ws2, row, "Tables",     best["n_tables"]);   row += 1
        kv(ws2, row, "Edges G_T",  best["n_edges"]);    row += 1
        kv(ws2, row, "theta_A",    best["theta_a"]);    row += 1
        kv(ws2, row, "theta_T",    best["theta_t"]);    row += 1
        kv(ws2, row, "Resolution", best["resolution"]); row += 2

    if failed:
        title_row(ws2, row, "Failed Runs"); row += 1
        for r in failed:
            kv(ws2, row, r["run_tag"], "FAILED"); row += 1

    wb.save(path)
    print(f"\nExcel report saved -> {path}")


def write_summary(results: list[dict], path: Path, dataset_dir: str) -> None:
    completed = [r for r in results if r["Q"] is not None]
    failed    = [r for r in results if r["status"] == "FAILED"]

    if not completed:
        path.write_text("No completed runs to summarise.\n", encoding="utf-8")
        return

    # Sort by Q descending
    completed.sort(key=lambda r: r["Q"], reverse=True)
    best = completed[0]

    lines = [
        "=" * 78,
        f"CCM Parameter Tuning Summary - {dataset_dir}",
        f"Generated : {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        f"Total runs: {len(results)}   Completed: {len(completed)}   Skipped: {len([r for r in results if r['status']=='skipped'])}   Failed: {len(failed)}",
        "=" * 78,
        "",
        f"  {'theta_A':>7}  {'theta_T':>7}  {'Res':>5}  {'Domains':>7}  {'Edges':>6}  {'Q':>7}  {'Time':>7}  Flag",
        "  " + "-" * 66,
    ]

    for r in completed:
        flag = "  < BEST" if r is best else ("  OK Q>0.3" if r["Q"] > 0.3 else "")
        lines.append(
            f"  {r['theta_a']:>7}  {r['theta_t']:>7}  {r['resolution']:>5}  "
            f"{str(r['n_domains']):>7}  {str(r['n_edges'] or '?'):>6}  "
            f"{r['Q']:>7.4f}  {str(r['elapsed_s'])+'s':>7}{flag}"
        )

    if failed:
        lines += ["", "Failed runs:"]
        for r in failed:
            lines.append(f"  {r['run_tag']}")

    lines += [
        "",
        f"Best run : {best['run_tag']}",
        f"  Q          = {best['Q']}",
        f"  Domains    = {best['n_domains']}",
        f"  Edges G_T  = {best['n_edges']}",
        f"  theta_A    = {best['theta_a']}",
        f"  theta_T    = {best['theta_t']}",
        f"  resolution = {best['resolution']}",
        "=" * 78,
    ]

    text = "\n".join(lines)
    path.write_text(text, encoding="utf-8")
    print("\n" + text)


# =============================================================================
# Main
# =============================================================================

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Parameter tuning over theta_A, theta_T, resolution for CCM pipeline.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )

    # Dataset
    parser.add_argument("--dataset_dir", required=True,
                        help="Dataset subfolder (e.g. adventure_works)")
    parser.add_argument("--schema",    default="schema.json")
    parser.add_argument("--knowledge", default="knowledge.docx",
                        help="Knowledge file (default: knowledge.docx). "
                             "Pass empty string to skip.")
    parser.add_argument("--out_dir",   default="ccm_output")

    # Grid
    parser.add_argument("--theta_a",    type=float, nargs="+",
                        default=DEFAULT_THETA_A,
                        help=f"theta_A values (default: {DEFAULT_THETA_A})")
    parser.add_argument("--theta_t",    type=float, nargs="+",
                        default=DEFAULT_THETA_T,
                        help=f"theta_T values (default: {DEFAULT_THETA_T})")
    parser.add_argument("--resolution", type=float, nargs="+",
                        default=DEFAULT_RESOLUTION,
                        help=f"resolution values (default: {DEFAULT_RESOLUTION})")

    # Pipeline passthrough
    parser.add_argument("--model",      default="mistral:latest")
    parser.add_argument("--ollama_url", default="http://localhost:11434")
    parser.add_argument("--no_llm",     action="store_true",
                        help="Skip LLM labeling - Q-only search (fast)")
    parser.add_argument("--start_from", default=None,
                        choices=["step12","step3a","step3b","step3c","step4","step5"],
                        help="Resume all runs from this step (default: step3c)")

    # Control
    parser.add_argument("--force",      action="store_true",
                        help="Re-run all combinations even if already complete")
    parser.add_argument("--dry_run",    action="store_true",
                        help="Print commands without executing")

    args = parser.parse_args()

    # Resolve script_dir — tune_params.py and run_pipeline.py are co-located
    project_root = PROJECT_ROOT
    dataset_path = resolve_dataset_dir(args.dataset_dir)
    dataset_name = dataset_display_name(args.dataset_dir, dataset_path)
    dataset_arg = str(dataset_path)

    # Resolve knowledge path
    knowledge = args.knowledge if args.knowledge else None
    if knowledge:
        k_path = dataset_path / knowledge
        if not k_path.exists():
            print(f"[WARN] knowledge file not found: {k_path} - continuing without it")
            knowledge = None

    # Default start_from: skip steps 1-3b if shared outputs already exist
    start_from = args.start_from
    if start_from is None:
        shared_ok = all(
            (dataset_path / args.out_dir / f).exists()
            for f in ["step1_concepts.json",
                      "step2_column_profiles.json",
                      "phi_matrix.csv",
                      "derived_weights.csv"]
        )
        start_from = "step3c" if shared_ok else None
        if shared_ok:
            print("\n[tune_params] Shared outputs detected - starting all runs from step3c.")
        else:
            print("\n[tune_params] Shared outputs not found - running full pipeline for each combination.")
            print("             Consider running once manually first to generate shared outputs.")

    # Build grid
    grid = list(product(args.theta_a, args.theta_t, args.resolution))
    total = len(grid)
    print(f"\n[tune_params] Dataset  : {dataset_name}")
    print(f"[tune_params] Path     : {dataset_path}")
    print(f"[tune_params] Grid size: {total} combinations")
    print(f"[tune_params] theta_A  : {args.theta_a}")
    print(f"[tune_params] theta_T  : {args.theta_t}")
    print(f"[tune_params] resolution: {args.resolution}")
    print(f"[tune_params] no_llm   : {args.no_llm}")
    print(f"[tune_params] start_from: {start_from or 'beginning'}")

    # Run grid
    results = []
    t_total = time.time()
    for i, (ta, tt, res) in enumerate(grid, 1):
        print(f"\n[tune_params] {i}/{total}  tA={ta}  tT={tt}  r={res}")
        r = run_one(
            script_dir   = project_root,
            dataset_dir  = dataset_arg,
            schema       = args.schema,
            knowledge    = knowledge,
            theta_a      = ta,
            theta_t      = tt,
            resolution   = res,
            model        = args.model,
            ollama_url   = args.ollama_url,
            no_llm       = args.no_llm,
            out_dir      = args.out_dir,
            dry_run      = args.dry_run,
            start_from   = start_from,
            force        = args.force,
        )
        results.append(r)

    elapsed_total = round(time.time() - t_total, 1)
    print(f"\n[tune_params] All {total} runs complete in {elapsed_total}s "
          f"({elapsed_total/60:.1f} min)")

    # Save outputs next to this script (or in dataset_dir)
    out_base = dataset_path / args.out_dir
    out_base.mkdir(parents=True, exist_ok=True)
    write_excel(results, out_base / "tune_params_results.xlsx", dataset_name)
    write_summary(results, out_base / "tune_params_summary.txt", dataset_name)


if __name__ == "__main__":
    main()
